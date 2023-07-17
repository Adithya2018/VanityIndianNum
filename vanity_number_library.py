import re
import openpyxl as opxl
from vanity_indian_no import VanityIndianNo


if __name__ == '__main__':
    
    # Taking the rule book for Vanity Check as input
    vanity_check_level_cost_wb = opxl.load_workbook("Data/Vanity_Check_bsnl_cost.xlsx")
    vanity_check_level_cost_sheet_names = vanity_check_level_cost_wb.sheetnames
    vanity_check_level_cost_ws = vanity_check_level_cost_wb[vanity_check_level_cost_sheet_names[0]]

    # Taking phone numbers as input from workbook
    vanity_check_wb = opxl.load_workbook("Data/vanity check.xlsx")
    vanity_check_sheet_names = vanity_check_wb.sheetnames
    vanity_check_ws = vanity_check_wb[vanity_check_sheet_names[0]]
    
    for i in range(2, vanity_check_ws.max_row):
        
        input_phone_no = vanity_check_ws.cell(row = i, column = 1)
        vn = VanityIndianNo(str(input_phone_no.value))

        if vn.is_valid_no(vn.phone_no):
            level, rate = vn.pattern_check(vn.phone_no)
            print(level, rate)
        else:
            print("Enter a valid Number")
